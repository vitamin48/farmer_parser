"""
Скрипт считывает файл JSON с товарами Фермера и записывает данные в Excel.
"""
import json
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE_NAME_JSON = 'data.json'  # out/FILE_NAME_JSON


def read_json():
    with open(f'out/{FILE_NAME_JSON}', 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        return data


def transform_price(x):
    x = int(round(float(x)))
    result = x * 5 if x < 200 else (
        x * 4.5 if 200 <= x < 500 else (
            x * 4 if 500 <= x < 1000 else (
                x * 3.5 if 1000 <= x < 5000 else (
                    x * 3 if 5000 <= x < 10000 else (
                        x * 2.5 if 10000 <= x < 20000 else (x * 2))))))
    # Убеждаемся, что значение после преобразований не меньше 490
    result = max(result, 490)
    # Округление до целого числа
    return round(result)


def create_df_by_dict(data_dict):
    """Создание DF из словаря"""
    pd.options.mode.copy_on_write = True
    """Из-за вложенности словарей преобразуем словарь к списку (df_list)"""
    df_list = []
    for art_id, data in data_dict.items():
        "Удаляем блок характеристики из data и превращаем его в отдельный словарь"
        characteristics = data.pop("characteristics", {})
        "Создаем словарь с Артикулами (row_data)"
        row_data = {"Артикул": art_id}
        "Обновляем словарь row_data другим словарем data, в котором все, кроме артикула и характеристик"
        row_data.update(data)
        "Обновляем словарь характеристиками (characteristics)"
        row_data.update(characteristics)
        "Добавляем результирующий словарь в список (df_list)"
        df_list.append(row_data)
    "Создаем DF на основе списка df_list"
    df = pd.DataFrame(df_list)
    "Создаем столбец с доп. изображениями"
    df['img_url2'] = df['img_url']  # Копируем столбец с изображениями
    # Переводим список с доп. изображениями начиная со 2 элемента в строку с разделителем (,)
    df['img_url2'] = df.apply(lambda row: ', '.join(row['img_url'][1:]) if len(row['img_url']) > 1 else '-', axis=1)
    "Создаем столбец с главным фото (только первое значение)"
    df['img_url'] = df['img_url'].apply(lambda x: x[0] if x else None)
    "Дабавляем в Артикул префикс fm_"
    df['Артикул'] = df['Артикул'].apply(lambda x: 'fm_' + x)
    "Работа с ценой"
    # # Создайте DataFrame df2, куда будут скопированы строки с пустыми значениями в "price" (пустые строки '')
    # df2 = df[df['price'] == ''].copy()
    # # Удаляем из оригинального df значения df2, которые могут содержать ошибки
    # df = df[~df.isin(df2)].dropna()
    df['price'] = df['price'].apply(lambda x: 0 if x == '' else x)
    # Создаем столбец Цена с учетом мин. заказа (price2)
    df['price2'] = df.apply(lambda row: int(round(float(row['price']) * float(row['Минимальный заказ:']))), axis=1)
    df['price3'] = df['price'].apply(transform_price)
    df['price4'] = df['price2'].apply(transform_price)
    # создаем столбец Брэнд со значениями NoName
    df['Брэнд'] = 'NoName'
    df['Страна'] = 'Россия'
    df['Вес:'] = df['Вес:'].apply(lambda x: int(round(float(x) * 1000)))
    "Формируем нужный порядок столбцов"
    df.insert(3, 'Минимальный заказ:', df.pop('Минимальный заказ:'))
    df.insert(4, 'price2', df.pop('price2'))
    df.insert(5, 'price3', df.pop('price3'))
    df.insert(6, 'price4', df.pop('price4'))
    df.insert(7, 'stock', df.pop('stock'))
    df.insert(8, 'Брэнд', df.pop('Брэнд'))
    df.insert(9, 'description', df.pop('description'))
    df.insert(10, 'Вес:', df.pop('Вес:'))
    df.insert(11, 'Ширина:', df.pop('Ширина:'))
    df.insert(12, 'Высота:', df.pop('Высота:'))
    df.insert(13, 'Длина:', df.pop('Длина:'))
    df.insert(14, 'Страна', df.pop('Страна'))
    df.insert(15, 'img_url', df.pop('img_url'))
    df.insert(16, 'img_url2', df.pop('img_url2'))
    df.insert(17, 'Объем:', df.pop('Объем:'))
    df.insert(18, 'Количество в упаковке:', df.pop('Количество в упаковке:'))
    df.insert(19, 'art_url', df.pop('art_url'))
    df.pop('Количество на паллете:')
    "Переименуем столбцы"
    df.columns = ['Артикул', 'Название', 'Цена закупа 1 ед.', 'Мин.заказ', 'Цена за мин.заказ',
                  'Цена для продажи 1 ед.', 'Цена для продажи мин.заказа', 'Остаток', 'Брэнд', 'Описание', 'Вес, г',
                  'Ширина, мм', 'Высота, мм', 'Длина, мм', 'Страна', 'Ссылка на главное фото товара',
                  'Ссылки на фото товара', 'Объем (м3)', 'Количество в упаковке', 'Ссылка на товар']

    new_df = df.head(100).copy()
    res_df_wb = df.copy()
    res_df_wb = res_df_wb.rename(columns={'Ширина, мм': 'Ширина, cм'})
    res_df_wb = res_df_wb.rename(columns={'Высота, мм': 'Высота, cм'})
    res_df_wb = res_df_wb.rename(columns={'Длина, мм': 'Длина, cм'})
    # Конверт мм в см
    res_df_wb['Ширина, cм'] = res_df_wb['Ширина, cм'].apply(
        lambda x: int(round(float(x) / 10)) if x != '-' else '-')
    res_df_wb['Высота, cм'] = res_df_wb['Высота, cм'].apply(
        lambda x: int(round(float(x) / 10)) if x != '-' else '-')
    res_df_wb['Длина, cм'] = res_df_wb['Длина, cм'].apply(
        lambda x: int(round(float(x) / 10)) if x != '-' else '-')
    # Сбросьте индекс для чистоты
    res_df_wb.reset_index(drop=True, inplace=True)
    return df, res_df_wb


# Функция для применения стилей к первой строке
def highlight_row(row):
    return ['background-color: red' if row['Цена закупа 1 ед.'] == 0 else
            'background-color: yellow' if row['Вес, г'] == 0 else '' for _ in row]


def create_xls(res_df_ozon, res_df_wb):
    file_name = f'out\\Farmer_big.xlsx'
    # Создание объекта Styler
    styled_df_ozon = res_df_ozon.style.apply(highlight_row, axis=1)
    styled_df_wb = res_df_wb.style.apply(highlight_row, axis=1)
    # Сохранение DataFrame в Excel с использованием Styler
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        styled_df_ozon.to_excel(writer, sheet_name='OZON', index=False, na_rep='NaN')
        styled_df_wb.to_excel(writer, sheet_name='WB', index=False, na_rep='NaN')

        # Установка ширины столбцов
        worksheet_ozon = writer.sheets['OZON']

        for column in res_df_ozon:
            column_width = max(res_df_ozon[column].astype(str).map(len).max(), len(column)) + 2
            col_letter = get_column_letter(res_df_ozon.columns.get_loc(column) + 1)
            worksheet_ozon.column_dimensions[col_letter].width = column_width

        # Закрепите первую строку
        worksheet_ozon.freeze_panes = 'A2'

        # Корректировка ширины столбцов
        worksheet_ozon.column_dimensions[get_column_letter(styled_df_ozon.columns.get_loc('Название') + 1)].width = 30
        worksheet_ozon.column_dimensions[get_column_letter(styled_df_ozon.columns.get_loc('Описание') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылка на главное фото товара') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылки на фото товара') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылка на товар') + 1)].width = 20

        # Установка ширины столбцов
        worksheet_wb = writer.sheets['WB']

        for column in res_df_ozon:
            column_width = max(res_df_ozon[column].astype(str).map(len).max(), len(column)) + 2
            col_letter = get_column_letter(res_df_ozon.columns.get_loc(column) + 1)
            worksheet_wb.column_dimensions[col_letter].width = column_width

        # Закрепите первую строку
        worksheet_wb.freeze_panes = 'A2'

        # Корректировка ширины столбцов
        worksheet_wb.column_dimensions[get_column_letter(styled_df_ozon.columns.get_loc('Название') + 1)].width = 30
        worksheet_wb.column_dimensions[get_column_letter(styled_df_ozon.columns.get_loc('Описание') + 1)].width = 30
        worksheet_wb.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылка на главное фото товара') + 1)].width = 30
        worksheet_wb.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылки на фото товара') + 1)].width = 30
        worksheet_wb.column_dimensions[
            get_column_letter(styled_df_ozon.columns.get_loc('Ссылка на товар') + 1)].width = 20


if __name__ == '__main__':
    data_json = read_json()
    res_df_ozon, res_df_wb = create_df_by_dict(data_dict=data_json)
    create_xls(res_df_ozon, res_df_wb)
